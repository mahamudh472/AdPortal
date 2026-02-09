import logging
from celery import shared_task
from datetime import datetime
from io import BytesIO
from django.core.files.base import ContentFile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

logger = logging.getLogger(__name__)

# Map of metric key -> display header
METRIC_MAP = {
    'spend': ('Spend', 'spend'),
    'impressions': ('Impressions', 'impressions'),
    'clicks': ('Clicks', 'clicks'),
    'ctr': ('CTR', 'ctr'),
    'cpc': ('CPC', 'cpc'),
    'roas': ('ROAS', 'roas'),
}

# Base columns always included
BASE_HEADERS = [
    ('ID', 'id'),
    ('Platform', 'platform'),
    ('Account ID', 'account_id'),
    ('Campaign ID', 'campaign_id'),
    ('Campaign Name', 'campaign_name'),
    ('Adgroup ID', 'adgroup_id'),
    ('Date', 'date'),
]


@shared_task(bind=True, max_retries=3)
def generate_report_task(self, report_id):
    """Background task to generate an Excel report."""
    from .models import AnalysisDaily, Report
    from main.models import Organization
    from accounts.models import Notification

    try:
        report = Report.objects.select_related('organization').get(id=report_id)
        organization = report.organization
        included_platforms = [p.upper() for p in (report.included_platforms or [])]
        included_metrics = [m.lower() for m in (report.included_metrics or [])]

        # Build dynamic headers based on selected metrics
        headers = list(BASE_HEADERS)
        for metric_key in included_metrics:
            if metric_key in METRIC_MAP:
                display_name, _ = METRIC_MAP[metric_key]
                headers.append((display_name, metric_key))

        # If no metrics selected, include all
        if not included_metrics:
            for key, (display_name, _) in METRIC_MAP.items():
                headers.append((display_name, key))

        # Gather data per platform
        platform_data = {}
        all_data = []

        for platform_key in included_platforms or ['META', 'GOOGLE', 'TIKTOK']:
            integration = organization.integrations.filter(platform=platform_key).first()
            if integration:
                qs = AnalysisDaily.objects.filter(
                    platform=platform_key,
                    account_id=integration.ad_account_id,
                )
                
                # Apply date range filter if available
                if report.start_date and report.end_date:
                    qs = qs.filter(date__gte=report.start_date, date__lte=report.end_date)
                
                data = list(qs.values())
                if data:
                    platform_data[platform_key] = data
                    all_data.extend(data)

        # ── Build Excel workbook ──
        wb = Workbook()
        header_names = [h[0] for h in headers]
        header_keys = [h[1] for h in headers]

        def populate_sheet(sheet, data, sheet_name):
            sheet.title = sheet_name

            for col_num, header in enumerate(header_names, 1):
                cell = sheet.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

            for row_num, record in enumerate(data, 2):
                for col_num, key in enumerate(header_keys, 1):
                    value = record.get(key)
                    # Convert date objects to string for Excel compatibility
                    if hasattr(value, 'isoformat'):
                        value = value.isoformat()
                    sheet.cell(row=row_num, column=col_num, value=value)

            for col in sheet.columns:
                max_length = 0
                column_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                sheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # All Data sheet
        populate_sheet(wb.active, all_data, "All Data")

        # Per-platform sheets
        for platform_key, data in platform_data.items():
            ws = wb.create_sheet(platform_key)
            populate_sheet(ws, data, platform_key)

        # Save to file
        filename = f"{report.report_type.capitalize()}_report_{organization.name}_{report.start_date.strftime('%Y%m%d')}_{report.end_date.strftime('%Y%m%d')}.xlsx"
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        report.file.save(filename, ContentFile(excel_file.read()), save=False)
        report.name = filename
        report.status = Report.Status.COMPLETED
        report.save()

        # Create in-app notification for all org owners / admins
        from main.models import OrganizationMember
        members = OrganizationMember.objects.filter(
            organization=organization,
            role__in=['OWNER', 'ADMIN'],
            status='ACTIVE',
        ).select_related('user')

        notifications = [
            Notification(
                user=member.user,
                organization=organization,
                message=f"Your report \"{filename}\" has been generated successfully and is ready for download.",
            )
            for member in members
        ]
        Notification.objects.bulk_create(notifications)

        logger.info("Report %s generated successfully for org %s", report_id, organization.name)

    except Exception as exc:
        logger.exception("Report generation failed for report_id=%s", report_id)
        try:
            report = Report.objects.get(id=report_id)
            report.status = Report.Status.FAILED
            report.save(update_fields=['status'])
        except Report.DoesNotExist:
            pass
        raise self.retry(exc=exc, countdown=60)
