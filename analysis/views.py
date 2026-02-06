from rest_framework.generics import GenericAPIView, ListAPIView
from main.mixins import RequiredOrganizationIDMixin
from rest_framework.response import Response
from main.models import Organization
from .models import AnalysisDaily, Report
from django.http import HttpResponse
from django.core.files.base import ContentFile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime
from io import BytesIO
from .serializers import ReportSerializer
from accounts.permissions import IsOrganizationMember, IsAdminOrOwnerOfOrganization, IsRegularPlatformUser

class ReportListView(RequiredOrganizationIDMixin, ListAPIView):
    serializer_class = ReportSerializer
    permission_classes = [IsOrganizationMember | IsAdminOrOwnerOfOrganization | IsRegularPlatformUser]
    
    def get_queryset(self):
        org_id = self.get_org_id()
        organization = Organization.objects.filter(snowflake_id=org_id).first()
        if organization:
            return Report.objects.filter(organization=organization).order_by('-created_at')
        return Report.objects.none()

# TODO: Need to make the report generation asynchronous using Celery and Redis to avoid long processing times and potential timeouts. This will allow the API to return a response immediately while the report is being generated in the background. Once the report is ready, we can notify the user via email or in-app notification with a link to download the report.
class GenerateReportView(RequiredOrganizationIDMixin, GenericAPIView):
    permission_classes = [IsRegularPlatformUser, IsAdminOrOwnerOfOrganization]
    def post(self, request, *args, **kwargs):
        org_id = self.get_org_id()
        organization = Organization.objects.filter(snowflake_id=org_id).first()

        if not organization:
            return Response({"error": "Organization not found"}, status=404)

        meta_platform = organization.integrations.filter(platform='META').first()
        google_platform = organization.integrations.filter(platform='GOOGLE').first()
        tiktok_platform = organization.integrations.filter(platform='TIKTOK').first()

        # Collect all data
        all_data = []
        meta_data = []
        google_data = []
        tiktok_data = []

        if meta_platform:
            meta_analysis = AnalysisDaily.objects.filter(platform='META', account_id=meta_platform.ad_account_id)
            meta_data = list(meta_analysis.values())
            all_data.extend(meta_data)

        if google_platform:
            google_analysis = AnalysisDaily.objects.filter(platform='GOOGLE', account_id=google_platform.ad_account_id)
            google_data = list(google_analysis.values())
            all_data.extend(google_data)

        if tiktok_platform:
            tiktok_analysis = AnalysisDaily.objects.filter(platform='TIKTOK', account_id=tiktok_platform.ad_account_id)
            tiktok_data = list(tiktok_analysis.values())
            all_data.extend(tiktok_data)

        # Create Excel file
        wb = Workbook()
        
        # Define headers
        headers = ['ID', 'Platform', 'Account ID', 'Campaign ID', 'Campaign Name', 'Adgroup ID', 
                   'Date', 'Impressions', 'Clicks', 'Spend', 'CTR', 'CPC', 'ROAS']
        
        # Helper function to add data to sheet
        def populate_sheet(sheet, data, sheet_name):
            sheet.title = sheet_name
            
            # Add headers with styling
            for col_num, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            
            # Add data rows
            for row_num, record in enumerate(data, 2):
                sheet.cell(row=row_num, column=1, value=record.get('id'))
                sheet.cell(row=row_num, column=2, value=record.get('platform'))
                sheet.cell(row=row_num, column=3, value=record.get('account_id'))
                sheet.cell(row=row_num, column=4, value=record.get('campaign_id'))
                sheet.cell(row=row_num, column=5, value=record.get('campaign_name'))
                sheet.cell(row=row_num, column=6, value=record.get('adgroup_id'))
                sheet.cell(row=row_num, column=7, value=record.get('date'))
                sheet.cell(row=row_num, column=8, value=record.get('impressions'))
                sheet.cell(row=row_num, column=9, value=record.get('clicks'))
                sheet.cell(row=row_num, column=10, value=record.get('spend'))
                sheet.cell(row=row_num, column=11, value=record.get('ctr'))
                sheet.cell(row=row_num, column=12, value=record.get('cpc'))
                sheet.cell(row=row_num, column=13, value=record.get('roas'))
            
            # Auto-adjust column widths
            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column].width = adjusted_width
        
        # Create sheets
        # 1. All Data sheet
        populate_sheet(wb.active, all_data, "All Data")
        
        # 2. META sheet
        if meta_data:
            meta_sheet = wb.create_sheet("META")
            populate_sheet(meta_sheet, meta_data, "META")
        
        # 3. GOOGLE sheet
        if google_data:
            google_sheet = wb.create_sheet("GOOGLE")
            populate_sheet(google_sheet, google_data, "GOOGLE")
        
        # 4. TIKTOK sheet
        if tiktok_data:
            tiktok_sheet = wb.create_sheet("TIKTOK")
            populate_sheet(tiktok_sheet, tiktok_data, "TIKTOK")
        
        # Generate filename with timestamp
        filename = f"analysis_report_{organization.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Save workbook to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Create Report instance and save the file
        report = Report(
            organization=organization,
            name=filename,
            report_type='analysis'
        )
        # Save the report first to get an ID
        report.save()
        # Then save the file
        report.file.save(filename, ContentFile(excel_file.read()), save=True)
        
        # Build the full file URL
        file_url = request.build_absolute_uri(report.file.url)
        
        return Response({
            "message": "Report generated successfully",
            "report_id": report.id,
            "file_url": file_url,
            "filename": filename,
            "created_at": report.created_at
        }, status=200)